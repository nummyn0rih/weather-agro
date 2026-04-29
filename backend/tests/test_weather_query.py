"""Tests for the universal `/api/weather/daily` query endpoint.

The aggregation primitives are pure functions, so we exercise them directly
with hand-crafted rows. The endpoint itself is covered by a small TestClient
test that monkeypatches the service layer.
"""

from __future__ import annotations

from datetime import date
from typing import AsyncIterator

import pytest
from fastapi.testclient import TestClient

from app.api import weather as weather_api
from app.api.deps import get_current_user
from app.db.models import User
from app.db.session import get_db
from app.main import app
from app.services.weather.export import rows_to_csv, rows_to_xlsx
from app.services.weather.query import (
    _bucket_start,
    _heatmap_x,
    aggregate_buckets,
    build_cumulative,
    build_heatmap,
    build_stats,
    collapse_to_average,
)


def test_bucket_start_week_aligns_to_monday() -> None:
    # 2026-04-29 is a Wednesday → bucket should be Monday 2026-04-27.
    assert _bucket_start(date(2026, 4, 29), "week") == date(2026, 4, 27)


def test_bucket_start_month_and_year() -> None:
    assert _bucket_start(date(2026, 4, 29), "month") == date(2026, 4, 1)
    assert _bucket_start(date(2026, 4, 29), "year") == date(2026, 1, 1)


def test_bucket_start_meteorological_seasons() -> None:
    # Winter spans Dec-Feb and the bucket key is the December of the start year.
    assert _bucket_start(date(2026, 1, 15), "season") == date(2025, 12, 1)
    assert _bucket_start(date(2025, 12, 15), "season") == date(2025, 12, 1)
    assert _bucket_start(date(2026, 4, 1), "season") == date(2026, 3, 1)
    assert _bucket_start(date(2026, 7, 1), "season") == date(2026, 6, 1)
    assert _bucket_start(date(2026, 10, 1), "season") == date(2026, 9, 1)


def test_collapse_to_average_means_across_sources() -> None:
    rows = [
        {"time": date(2026, 4, 1), "location_id": 1, "source": "open_meteo",
         "temp_avg": 10.0, "precipitation": 2.0},
        {"time": date(2026, 4, 1), "location_id": 1, "source": "nasa_power",
         "temp_avg": 12.0, "precipitation": None},
        {"time": date(2026, 4, 2), "location_id": 1, "source": "open_meteo",
         "temp_avg": None, "precipitation": 5.0},
    ]
    out = collapse_to_average(rows, ["temp_avg", "precipitation"])
    by_day = {r["time"]: r for r in out}

    assert by_day[date(2026, 4, 1)]["source"] == "average"
    assert by_day[date(2026, 4, 1)]["temp_avg"] == pytest.approx(11.0)
    # Only one source had a precipitation value — average over present sources.
    assert by_day[date(2026, 4, 1)]["precipitation"] == pytest.approx(2.0)
    assert by_day[date(2026, 4, 2)]["temp_avg"] is None
    assert by_day[date(2026, 4, 2)]["precipitation"] == pytest.approx(5.0)


def test_aggregate_buckets_day_passthrough_sorted() -> None:
    rows = [
        {"time": date(2026, 4, 2), "location_id": 1, "source": "open_meteo", "temp_avg": 11.0},
        {"time": date(2026, 4, 1), "location_id": 1, "source": "open_meteo", "temp_avg": 10.0},
    ]
    out = aggregate_buckets(rows, ["temp_avg"], "day", "open_meteo")
    assert [r["time"] for r in out] == [date(2026, 4, 1), date(2026, 4, 2)]


def test_aggregate_buckets_month_averages_and_sums() -> None:
    rows = [
        {"time": date(2026, 4, 1), "location_id": 1, "source": "open_meteo",
         "temp_avg": 10.0, "precipitation": 1.0},
        {"time": date(2026, 4, 15), "location_id": 1, "source": "open_meteo",
         "temp_avg": 14.0, "precipitation": 2.5},
        {"time": date(2026, 5, 1), "location_id": 1, "source": "open_meteo",
         "temp_avg": 18.0, "precipitation": 0.0},
    ]
    out = aggregate_buckets(rows, ["temp_avg", "precipitation"], "month", "open_meteo")
    by_month = {r["time"]: r for r in out}

    # April: temp_avg averaged, precipitation summed.
    assert by_month[date(2026, 4, 1)]["temp_avg"] == pytest.approx(12.0)
    assert by_month[date(2026, 4, 1)]["precipitation"] == pytest.approx(3.5)
    # May: single sample.
    assert by_month[date(2026, 5, 1)]["temp_avg"] == pytest.approx(18.0)
    assert by_month[date(2026, 5, 1)]["precipitation"] == pytest.approx(0.0)


def test_aggregate_buckets_handles_all_none() -> None:
    rows = [
        {"time": date(2026, 4, 1), "location_id": 1, "source": "open_meteo",
         "temp_avg": None, "precipitation": None},
        {"time": date(2026, 4, 2), "location_id": 1, "source": "open_meteo",
         "temp_avg": None, "precipitation": None},
    ]
    out = aggregate_buckets(rows, ["temp_avg", "precipitation"], "month", "open_meteo")
    assert len(out) == 1
    assert out[0]["temp_avg"] is None
    assert out[0]["precipitation"] is None


def test_aggregate_buckets_separates_locations() -> None:
    rows = [
        {"time": date(2026, 4, 1), "location_id": 1, "source": "open_meteo", "temp_avg": 10.0},
        {"time": date(2026, 4, 1), "location_id": 2, "source": "open_meteo", "temp_avg": 20.0},
    ]
    out = aggregate_buckets(rows, ["temp_avg"], "month", "open_meteo")
    assert len(out) == 2
    by_loc = {r["location_id"]: r["temp_avg"] for r in out}
    assert by_loc[1] == pytest.approx(10.0)
    assert by_loc[2] == pytest.approx(20.0)


@pytest.fixture
def client(monkeypatch):
    captured: dict[str, object] = {}

    async def fake_query(_session, **kwargs):
        captured.update(kwargs)
        return [
            {
                "time": date(2026, 4, 1),
                "location_id": 1,
                "source": "average",
                "temp_avg": 11.0,
            }
        ]

    monkeypatch.setattr(weather_api.query_service, "query_daily", fake_query)

    async def fake_user() -> User:
        return User(id=1, username="admin", password_hash="x")

    async def fake_db() -> AsyncIterator[None]:
        yield None

    app.dependency_overrides[get_current_user] = fake_user
    app.dependency_overrides[get_db] = fake_db

    with TestClient(app) as c:
        yield c, captured

    app.dependency_overrides.clear()


def test_endpoint_passes_params_and_returns_rows(client) -> None:
    c, captured = client
    response = c.get(
        "/api/weather/daily",
        params=[
            ("location_ids", 1),
            ("location_ids", 2),
            ("parameters", "temp_avg"),
            ("parameters", "precipitation"),
            ("date_from", "2026-01-01"),
            ("date_to", "2026-04-30"),
            ("source", "average"),
            ("aggregation", "month"),
        ],
    )
    assert response.status_code == 200
    body = response.json()
    assert body == [
        {
            "time": "2026-04-01",
            "location_id": 1,
            "source": "average",
            "temp_avg": 11.0,
        }
    ]
    assert captured["location_ids"] == [1, 2]
    assert captured["parameters"] == ["temp_avg", "precipitation"]
    assert captured["date_from"] == date(2026, 1, 1)
    assert captured["date_to"] == date(2026, 4, 30)
    assert captured["source"] == "average"
    assert captured["aggregation"] == "month"


def test_endpoint_validates_source_enum(client) -> None:
    c, _ = client
    response = c.get(
        "/api/weather/daily",
        params=[
            ("location_ids", 1),
            ("parameters", "temp_avg"),
            ("date_from", "2026-01-01"),
            ("date_to", "2026-04-30"),
            ("source", "garbage"),
        ],
    )
    assert response.status_code == 422


def test_heatmap_x_axes() -> None:
    d = date(2026, 4, 29)
    assert _heatmap_x(d, "month") == 4
    assert _heatmap_x(d, "week") == d.isocalendar().week
    assert _heatmap_x(d, "doy") == d.timetuple().tm_yday


def test_build_heatmap_avg_and_sum() -> None:
    rows = [
        {"time": date(2025, 4, 1), "location_id": 1, "source": "open_meteo", "temp_avg": 10.0},
        {"time": date(2025, 4, 15), "location_id": 1, "source": "open_meteo", "temp_avg": 14.0},
        {"time": date(2026, 4, 1), "location_id": 1, "source": "open_meteo", "temp_avg": 8.0},
    ]
    cells = build_heatmap(rows, "temp_avg", "month", "open_meteo")
    by_year = {(c["year"], c["x"]): c for c in cells}
    assert by_year[(2025, 4)]["value"] == pytest.approx(12.0)
    assert by_year[(2026, 4)]["value"] == pytest.approx(8.0)

    rows_p = [
        {"time": date(2025, 4, 1), "location_id": 1, "source": "open_meteo", "precipitation": 1.0},
        {"time": date(2025, 4, 5), "location_id": 1, "source": "open_meteo", "precipitation": 2.5},
    ]
    cells_p = build_heatmap(rows_p, "precipitation", "month", "open_meteo")
    assert cells_p[0]["value"] == pytest.approx(3.5)


def test_build_cumulative_precipitation_running_sum() -> None:
    rows = [
        {"time": date(2026, 4, 2), "location_id": 1, "source": "open_meteo", "precipitation": 2.0},
        {"time": date(2026, 4, 1), "location_id": 1, "source": "open_meteo", "precipitation": 1.0},
        {"time": date(2026, 4, 3), "location_id": 1, "source": "open_meteo", "precipitation": None},
        {"time": date(2026, 4, 4), "location_id": 1, "source": "open_meteo", "precipitation": 0.5},
    ]
    out = build_cumulative(
        rows, parameter="precipitation", base_temperature=None, out_source="open_meteo"
    )
    assert [r["time"] for r in out] == [
        date(2026, 4, 1), date(2026, 4, 2), date(2026, 4, 3), date(2026, 4, 4)
    ]
    assert [r["cumulative"] for r in out] == pytest.approx([1.0, 3.0, 3.0, 3.5])
    assert out[2]["daily"] is None


def test_build_cumulative_gdd_uses_base_temperature() -> None:
    rows = [
        {"time": date(2026, 4, 1), "location_id": 1, "source": "open_meteo",
         "temp_min": 8.0, "temp_max": 16.0},  # mean=12, base=10 → 2
        {"time": date(2026, 4, 2), "location_id": 1, "source": "open_meteo",
         "temp_min": 5.0, "temp_max": 9.0},   # mean=7, base=10 → 0 (clamped)
        {"time": date(2026, 4, 3), "location_id": 1, "source": "open_meteo",
         "temp_min": None, "temp_max": 20.0}, # missing → None
    ]
    out = build_cumulative(
        rows, parameter="gdd", base_temperature=10.0, out_source="open_meteo"
    )
    assert [r["daily"] for r in out] == [pytest.approx(2.0), pytest.approx(0.0), None]
    assert [r["cumulative"] for r in out] == pytest.approx([2.0, 2.0, 2.0])


def test_build_cumulative_gdd_requires_base_temperature() -> None:
    with pytest.raises(ValueError, match="base_temperature"):
        build_cumulative(
            [{"time": date(2026, 4, 1), "location_id": 1, "source": "open_meteo",
              "temp_min": 5.0, "temp_max": 10.0}],
            parameter="gdd",
            base_temperature=None,
            out_source="open_meteo",
        )


def test_build_cumulative_separates_locations() -> None:
    rows = [
        {"time": date(2026, 4, 1), "location_id": 1, "source": "open_meteo", "precipitation": 1.0},
        {"time": date(2026, 4, 1), "location_id": 2, "source": "open_meteo", "precipitation": 5.0},
        {"time": date(2026, 4, 2), "location_id": 1, "source": "open_meteo", "precipitation": 1.0},
    ]
    out = build_cumulative(
        rows, parameter="precipitation", base_temperature=None, out_source="open_meteo"
    )
    by_loc = {(r["location_id"], r["time"]): r["cumulative"] for r in out}
    assert by_loc[(1, date(2026, 4, 1))] == pytest.approx(1.0)
    assert by_loc[(1, date(2026, 4, 2))] == pytest.approx(2.0)
    assert by_loc[(2, date(2026, 4, 1))] == pytest.approx(5.0)


def test_endpoint_compare_years_overlay(client) -> None:
    c, captured = client
    response = c.get(
        "/api/weather/daily",
        params=[
            ("location_ids", 1),
            ("parameters", "temp_avg"),
            ("date_from", "2026-04-01"),
            ("date_to", "2026-04-30"),
            ("compare_years", 2024),
            ("compare_years", 2025),
        ],
    )
    assert response.status_code == 200
    assert captured["compare_years"] == [2024, 2025]


def test_endpoint_heatmap_calls_service(monkeypatch, client) -> None:
    c, _ = client

    captured: dict[str, object] = {}

    async def fake_heatmap(_session, **kwargs):
        captured.update(kwargs)
        return [
            {"location_id": 1, "parameter": "temp_avg", "source": "open_meteo",
             "year": 2026, "x": 4, "value": 12.0}
        ]

    monkeypatch.setattr(weather_api.query_service, "query_heatmap", fake_heatmap)
    response = c.get(
        "/api/weather/heatmap",
        params={
            "location_id": 1,
            "parameter": "temp_avg",
            "date_from": "2024-01-01",
            "date_to": "2026-12-31",
            "axis": "month",
        },
    )
    assert response.status_code == 200
    assert captured["location_id"] == 1
    assert captured["parameter"] == "temp_avg"
    assert captured["axis"] == "month"
    assert response.json()[0]["value"] == 12.0


def test_endpoint_cumulative_gdd_requires_base_temp(monkeypatch, client) -> None:
    c, _ = client

    async def boom(*_args, **_kwargs):
        raise ValueError("base_temperature is required for parameter='gdd'")

    monkeypatch.setattr(weather_api.query_service, "query_cumulative", boom)
    response = c.get(
        "/api/weather/cumulative",
        params=[
            ("location_ids", 1),
            ("parameter", "gdd"),
            ("date_from", "2026-04-01"),
            ("date_to", "2026-05-01"),
        ],
    )
    assert response.status_code == 400
    assert "base_temperature" in response.json()["detail"]


def test_endpoint_cumulative_happy_path(monkeypatch, client) -> None:
    c, _ = client

    captured: dict[str, object] = {}

    async def fake_cum(_session, **kwargs):
        captured.update(kwargs)
        return [
            {"time": date(2026, 4, 1), "location_id": 1, "source": "open_meteo",
             "parameter": "precipitation", "daily": 1.0, "cumulative": 1.0}
        ]

    monkeypatch.setattr(weather_api.query_service, "query_cumulative", fake_cum)
    response = c.get(
        "/api/weather/cumulative",
        params=[
            ("location_ids", 1),
            ("parameter", "precipitation"),
            ("date_from", "2026-04-01"),
            ("date_to", "2026-04-30"),
        ],
    )
    assert response.status_code == 200
    assert captured["parameter"] == "precipitation"
    body = response.json()
    assert body[0]["cumulative"] == 1.0


def test_build_stats_min_max_mean_sum_count_by_month() -> None:
    rows = [
        {"time": date(2026, 4, 1), "location_id": 1, "source": "open_meteo",
         "temp_avg": 10.0, "precipitation": 1.0},
        {"time": date(2026, 4, 15), "location_id": 1, "source": "open_meteo",
         "temp_avg": 14.0, "precipitation": 2.5},
        {"time": date(2026, 5, 1), "location_id": 1, "source": "open_meteo",
         "temp_avg": 18.0, "precipitation": None},
    ]
    out = build_stats(rows, ["temp_avg", "precipitation"], "month", "open_meteo")
    by = {(r["time"], r["parameter"]): r for r in out}
    apr_t = by[(date(2026, 4, 1), "temp_avg")]
    assert apr_t["min"] == pytest.approx(10.0)
    assert apr_t["max"] == pytest.approx(14.0)
    assert apr_t["mean"] == pytest.approx(12.0)
    assert apr_t["sum"] == pytest.approx(24.0)
    assert apr_t["count"] == 2

    apr_p = by[(date(2026, 4, 1), "precipitation")]
    assert apr_p["sum"] == pytest.approx(3.5)
    assert apr_p["count"] == 2

    may_p = by[(date(2026, 5, 1), "precipitation")]
    assert may_p["min"] is None
    assert may_p["max"] is None
    assert may_p["mean"] is None
    assert may_p["sum"] is None
    assert may_p["count"] == 0


def test_build_stats_total_collapses_range_to_one_bucket() -> None:
    rows = [
        {"time": date(2026, 1, 5), "location_id": 1, "source": "open_meteo", "temp_avg": 0.0},
        {"time": date(2026, 4, 15), "location_id": 1, "source": "open_meteo", "temp_avg": 10.0},
        {"time": date(2026, 7, 1), "location_id": 1, "source": "open_meteo", "temp_avg": 20.0},
    ]
    out = build_stats(rows, ["temp_avg"], "total", "open_meteo")
    assert len(out) == 1
    r = out[0]
    assert r["time"] == date(2026, 1, 5)
    assert r["min"] == pytest.approx(0.0)
    assert r["max"] == pytest.approx(20.0)
    assert r["mean"] == pytest.approx(10.0)
    assert r["count"] == 3


def test_endpoint_stats_passes_filters(monkeypatch, client) -> None:
    c, _ = client
    captured: dict[str, object] = {}

    async def fake_stats(_session, **kwargs):
        captured.update(kwargs)
        return [
            {"time": date(2026, 4, 1), "location_id": 1, "source": "open_meteo",
             "parameter": "temp_avg", "min": 5.0, "max": 15.0, "mean": 10.0,
             "sum": 30.0, "count": 3}
        ]

    monkeypatch.setattr(weather_api.query_service, "query_stats", fake_stats)
    response = c.get(
        "/api/weather/stats",
        params=[
            ("location_ids", 1),
            ("location_ids", 2),
            ("parameters", "temp_avg"),
            ("date_from", "2026-01-01"),
            ("date_to", "2026-04-30"),
            ("source", "average"),
            ("aggregation", "month"),
        ],
    )
    assert response.status_code == 200
    assert captured["location_ids"] == [1, 2]
    assert captured["parameters"] == ["temp_avg"]
    assert captured["aggregation"] == "month"
    body = response.json()
    assert body[0]["mean"] == 10.0
    assert body[0]["count"] == 3


def test_rows_to_csv_has_header_and_iso_dates() -> None:
    rows = [
        {"time": date(2026, 4, 1), "location_id": 1, "source": "open_meteo",
         "temp_avg": 10.0, "precipitation": 1.5},
        {"time": date(2026, 4, 2), "location_id": 1, "source": "open_meteo",
         "temp_avg": None, "precipitation": 0.0},
    ]
    out = rows_to_csv(rows, ["temp_avg", "precipitation"])
    lines = out.strip().split("\n")
    assert lines[0] == "time,location_id,source,temp_avg,precipitation"
    assert lines[1] == "2026-04-01,1,open_meteo,10.0,1.5"
    # None renders as empty cell.
    assert lines[2] == "2026-04-02,1,open_meteo,,0.0"


def test_rows_to_xlsx_returns_valid_workbook() -> None:
    import io

    from openpyxl import load_workbook

    rows = [
        {"time": date(2026, 4, 1), "location_id": 1, "source": "open_meteo",
         "temp_avg": 11.0, "precipitation": 2.0},
    ]
    blob = rows_to_xlsx(rows, ["temp_avg", "precipitation"])
    wb = load_workbook(io.BytesIO(blob), read_only=True)
    ws = wb["weather"]
    grid = [list(r) for r in ws.iter_rows(values_only=True)]
    assert grid[0] == ["time", "location_id", "source", "temp_avg", "precipitation"]
    assert grid[1] == ["2026-04-01", 1, "open_meteo", 11.0, 2.0]


def test_endpoint_export_csv(monkeypatch, client) -> None:
    c, _ = client

    async def fake_query(_session, **_kwargs):
        return [
            {"time": date(2026, 4, 1), "location_id": 1, "source": "open_meteo",
             "temp_avg": 10.0}
        ]

    monkeypatch.setattr(weather_api.query_service, "query_daily", fake_query)
    response = c.get(
        "/api/weather/export",
        params=[
            ("location_ids", 1),
            ("parameters", "temp_avg"),
            ("date_from", "2026-04-01"),
            ("date_to", "2026-04-30"),
            ("format", "csv"),
        ],
    )
    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/csv")
    assert "attachment" in response.headers["content-disposition"]
    body = response.text
    assert body.startswith("time,location_id,source,temp_avg")
    assert "2026-04-01,1,open_meteo,10.0" in body


def test_endpoint_export_xlsx(monkeypatch, client) -> None:
    c, _ = client

    async def fake_query(_session, **_kwargs):
        return [
            {"time": date(2026, 4, 1), "location_id": 1, "source": "open_meteo",
             "temp_avg": 10.0}
        ]

    monkeypatch.setattr(weather_api.query_service, "query_daily", fake_query)
    response = c.get(
        "/api/weather/export",
        params=[
            ("location_ids", 1),
            ("parameters", "temp_avg"),
            ("date_from", "2026-04-01"),
            ("date_to", "2026-04-30"),
            ("format", "xlsx"),
        ],
    )
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    # XLSX is a zip — magic bytes 'PK'.
    assert response.content[:2] == b"PK"


def test_endpoint_export_validates_format(monkeypatch, client) -> None:
    c, _ = client

    async def fake_query(_session, **_kwargs):
        return []

    monkeypatch.setattr(weather_api.query_service, "query_daily", fake_query)
    response = c.get(
        "/api/weather/export",
        params=[
            ("location_ids", 1),
            ("parameters", "temp_avg"),
            ("date_from", "2026-04-01"),
            ("date_to", "2026-04-30"),
            ("format", "pdf"),
        ],
    )
    assert response.status_code == 422


def test_endpoint_propagates_value_error_as_400(client, monkeypatch) -> None:
    c, _ = client

    async def boom(*_args, **_kwargs):
        raise ValueError("Unknown parameters: ['nope']")

    monkeypatch.setattr(weather_api.query_service, "query_daily", boom)
    response = c.get(
        "/api/weather/daily",
        params=[
            ("location_ids", 1),
            ("parameters", "nope"),
            ("date_from", "2026-01-01"),
            ("date_to", "2026-04-30"),
        ],
    )
    assert response.status_code == 400
    assert "Unknown parameters" in response.json()["detail"]
